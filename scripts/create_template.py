# ═══════════════════════════════════════════════════════════════
# Create Excel Template with 4 sheets
# ═══════════════════════════════════════════════════════════════

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from pathlib import Path

def create_template(filename: str = 'upload_tracker.xlsx'):
    """Create Excel template with 4 sheets"""
    
    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # ═══════════════════════════════════════════════════════════════
    # SHEET 1: VIDEOS (Main sheet)
    # ═══════════════════════════════════════════════════════════════
    ws_videos = wb.create_sheet('Videos', 0)
    
    videos_headers = [
        'Video File', 'Title', 'Caption', 'Description', 'Tags',
        'Publish Date', 'Status', 'IG URL', 'YT URL', 'TT URL', 'Error'
    ]
    
    for col, header in enumerate(videos_headers, 1):
        cell = ws_videos.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Sample data
    sample_data = [
        [
            'video1.mp4',
            'Amazing Place - 5 Years Left',
            'This place will disappear in 5 years\n\nWhile you\'re reading this — nobody\'s there...',
            'Cinematic travel video from Bali',
            'travel,viral,nature,bali,cinematic',
            f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}',
            'scheduled',
            '',
            '',
            '',
            ''
        ]
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_videos.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Set column widths
    ws_videos.column_dimensions['A'].width = 15
    ws_videos.column_dimensions['B'].width = 25
    ws_videos.column_dimensions['C'].width = 40
    ws_videos.column_dimensions['D'].width = 30
    ws_videos.column_dimensions['E'].width = 25
    ws_videos.column_dimensions['F'].width = 20
    ws_videos.column_dimensions['G'].width = 12
    ws_videos.column_dimensions['H'].width = 30
    ws_videos.column_dimensions['I'].width = 30
    ws_videos.column_dimensions['J'].width = 30
    ws_videos.column_dimensions['K'].width = 25
    
    # Set row height
    ws_videos.row_dimensions[1].height = 30
    ws_videos.row_dimensions[2].height = 40
    
    # ═══════════════════════════════════════════════════════════════
    # SHEET 2: INSTAGRAM LOGS
    # ═══════════════════════════════════════════════════════════════
    ws_ig = wb.create_sheet('Instagram Logs', 1)
    
    ig_headers = [
        'Timestamp', 'Video File', 'Status', 'URL', 'Error', 'Duration (sec)'
    ]
    
    for col, header in enumerate(ig_headers, 1):
        cell = ws_ig.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill(start_color="E1306C", end_color="E1306C", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    ws_ig.column_dimensions['A'].width = 20
    ws_ig.column_dimensions['B'].width = 15
    ws_ig.column_dimensions['C'].width = 12
    ws_ig.column_dimensions['D'].width = 40
    ws_ig.column_dimensions['E'].width = 30
    ws_ig.column_dimensions['F'].width = 15
    ws_ig.row_dimensions[1].height = 25
    
    # ═══════════════════════════════════════════════════════════════
    # SHEET 3: YOUTUBE LOGS
    # ═══════════════════════════════════════════════════════════════
    ws_yt = wb.create_sheet('YouTube Logs', 2)
    
    yt_headers = [
        'Timestamp', 'Video File', 'Status', 'Video ID', 'URL', 'Error', 'Duration (sec)'
    ]
    
    for col, header in enumerate(yt_headers, 1):
        cell = ws_yt.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    ws_yt.column_dimensions['A'].width = 20
    ws_yt.column_dimensions['B'].width = 15
    ws_yt.column_dimensions['C'].width = 12
    ws_yt.column_dimensions['D'].width = 15
    ws_yt.column_dimensions['E'].width = 40
    ws_yt.column_dimensions['F'].width = 30
    ws_yt.column_dimensions['G'].width = 15
    ws_yt.row_dimensions[1].height = 25
    
    # ═══════════════════════════════════════════════════════════════
    # SHEET 4: TIKTOK LOGS
    # ═══════════════════════════════════════════════════════════════
    ws_tt = wb.create_sheet('TikTok Logs', 3)
    
    tt_headers = [
        'Timestamp', 'Video File', 'Status', 'URL', 'Error', 'Duration (sec)'
    ]
    
    for col, header in enumerate(tt_headers, 1):
        cell = ws_tt.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    ws_tt.column_dimensions['A'].width = 20
    ws_tt.column_dimensions['B'].width = 15
    ws_tt.column_dimensions['C'].width = 12
    ws_tt.column_dimensions['D'].width = 40
    ws_tt.column_dimensions['E'].width = 30
    ws_tt.column_dimensions['F'].width = 15
    ws_tt.row_dimensions[1].height = 25
    
    # Save workbook
    wb.save(filename)
    print(f"✅ Created: {filename}")
    print(f"   - Sheet 1: Videos (main sheet with all video data)")
    print(f"   - Sheet 2: Instagram Logs (upload history)")
    print(f"   - Sheet 3: YouTube Logs (upload history)")
    print(f"   - Sheet 4: TikTok Logs (upload history)")
    print(f"\nColumns in Videos sheet:")
    for i, header in enumerate(videos_headers, 1):
        print(f"   {i}. {header}")

if __name__ == "__main__":
    import sys
    
    filename = sys.argv[1] if len(sys.argv) > 1 else 'upload_tracker.xlsx'
    create_template(filename)
